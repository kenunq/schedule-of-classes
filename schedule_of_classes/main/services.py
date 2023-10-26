#подключаем необходимы библиотеки
import os
import threading
import time
import shutil
import requests
from urllib.parse import urlencode
from urllib.request import urlopen
from zipfile import ZipFile

from openpyxl import load_workbook


def get_json(filename, find):
    '''Функция для создания JSON объекта путем парсинга excel таблицы'''
    book = load_workbook(filename='E:/schedule-of-classes/schedule_of_classes/static/other/Pасписание/' + filename)
    ws = book.active
    # namep = (os.path.basename(filename))
    masstime = []
    masssur = []
    massgr = []
    masskb = []
    find = find
    for i in range(1, 18):
        for j in range(2, 34):
            lol = ws.cell(row=j, column=i).value
            if lol is None:
                lol = '@'
            if isinstance(lol, str):
                if find in lol:
                    if 3 <= j <= 8:
                        qaz = ws.cell(row=2, column=i).value
                    elif 10 <= j <= 15:
                        qaz = ws.cell(row=9, column=i).value
                    elif 17 <= j <= 22:
                        qaz = ws.cell(row=16, column=i).value
                    elif 24 <= j <= 29:
                        qaz = ws.cell(row=23, column=i).value
                    masstime.insert(0, ws.cell(row=j, column=2).value)
                    masssur.insert(0, lol.strip().replace('\n', ' '))
                    massgr.insert(0, qaz)
                    masskb.insert(0, ws.cell(row=j, column=i + 1).value)

    lenght = len(masstime)

    pre_result = {}
    #заполняем словарь данными
    for i in range(lenght):
        pre_result[i] = {'Время: ': str(masstime[i]), 'Предмет и преподаватель: ': str(masssur[i]),
                         'Группа: ': str(massgr[i]), 'Кабинет: ': str(masskb[i])}

    #сортируем спарсенные данные по времени
    sotr_dict_key = sorted(pre_result, key=lambda x: int(pre_result[x]['Время: '].split('.')[0]))

    #перезаписываем отсортированные данные в словарь
    result = {}
    for i in range(lenght):
        result[i] = {'Время: ': str(masstime[sotr_dict_key[i]]), 'Предмет и преподаватель: ': str(masssur[sotr_dict_key[i]]),
                     'Группа: ': str(massgr[sotr_dict_key[i]]), 'Кабинет: ': str(masskb[sotr_dict_key[i]])}

    return result


def get_filename_json():
    '''Функция для получения доступных рассписаний занятий'''
    list_name = os.listdir('E:\schedule-of-classes\schedule_of_classes\static\other\Pасписание')

    result = {}
    for i in range(len(list_name)):
        result[i] = list_name[i]

    return result


def download_exel(url):
    '''Функция для скачивания excel таблиц с рассписанием'''
    base_url = 'https://cloud-api.yandex.net/v1/disk/public/resources/download?'
    public_key = url  # Сюда вписываете вашу ссылку

    # Получаем загрузочную ссылку
    final_url = base_url + urlencode(dict(public_key=public_key))
    response = requests.get(final_url)
    download_url = response.json()['href']
    to_url = os.path.dirname(os.path.abspath(__file__))[:-5]
    print(to_url)
    # записываем в переменную бинарные данные
    filedata = urlopen(download_url)
    datatowrite = filedata.read()

    # переводим бинарные данные в зип файл
    with open(to_url + '\static\other\qwe.zip', 'wb') as f:
        f.write(datatowrite)

    # разархивируем архив
    with ZipFile(to_url + '\static\other\qwe.zip', 'r') as zip_file:
        zip_file.extractall(to_url + '\static\other')


def del_from_folder():
    '''Функция для удаления папки с рассписаниями'''
    try:
        list = os.listdir("E:\schedule-of-classes\schedule_of_classes\static\other\Pасписание")
        if list:
            shutil.rmtree("E:\schedule-of-classes\schedule_of_classes\static\other\Pасписание")
    except:
        print("Folder not found")



class AsyncActionGetGameChatData(threading.Thread):
    '''Класс работающий не в основном потоке, который качает excel таблиц с рассписанием'''

    def run(self):
        while True:
            del_from_folder()
            download_exel('https://disk.yandex.ru/d/VNdnX6hmveqJuw')
            print('Скачивание завершилось')
            time.sleep(7200)


#создаем экземпляр класса для скачивания excel таблиц
async_action_get_game_chat_data = AsyncActionGetGameChatData()
async_action_get_game_chat_data.start()

# get_filename_json()

